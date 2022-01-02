import openpyxl
import os
from pathlib import Path
import shutil, errno  
import re

# load excel with its path
wrkbk = openpyxl.load_workbook(r".xlsx file.....")
ws = wrkbk.get_sheet_by_name('sheet name')  
sh = wrkbk.active


# iterate through excel and display data
for row in sh.iter_rows(min_row=1, min_col=1, max_row=12, max_col=3):
    for cell in row:
        for cell in ws['A']:
            acc_number_pre = (cell.value)
            temp1 = acc_number_pre.replace('files', '')
            acc_number1 = re.split("/", temp1)
            acc_num_acc = acc_number1[3]
            acc_number2 = acc_num_acc.replace('.txt', '')
            patient_ID = os.path.join(acc_number1[1] + '/' + acc_number1[2] + '/' + acc_number2 + '/')
            apex_url = "wget -r -c -np -N --no-check-certificate --user ...name.. --password ...password... -o file_download.log --progress=bar:force:noscroll "
            pre_path = (r" https://url......")
            full_url = (apex_url + pre_path + patient_ID)
            print(full_url)
            os.system(full_url)





